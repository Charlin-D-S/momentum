"""
Affichage et export du tableau des seuils (une ligne par niveau de score).

    afficher_tableau(st, t, s)   -> st.dataframe formaté + boutons CSV / Excel
    table_lisible(t)             -> la table avec des noms de colonnes parlants
    vers_csv(t) / vers_excel(t)  -> bytes prêts pour st.download_button
"""

from io import BytesIO

import pandas as pd
import polars as pl

COLONNES = {
    "proba": "Seuil (proba)",
    "n": "Dossiers au niveau",
    "n_defaut": "Défauts au niveau",
    "n_cum": "Acceptés (cumul)",
    "defaut_cum": "Défauts acceptés (cumul)",
    "taux_acceptation": "Part d'acceptés",
    "taux_defaut": "Taux de défaut des acceptés",
}


def table_lisible(t: pl.DataFrame, en_pourcent: bool = True,
                  lisse: bool = True) -> pl.DataFrame:
    """Renomme et met en forme la table produite par table_seuils()."""
    out = t.with_columns([                      # effectifs en entiers
        pl.col(c).round().cast(pl.Int64)
        for c in ("n", "n_defaut", "n_cum", "defaut_cum") if c in t.columns
    ])
    if lisse:
        out = out.with_columns(
            taux_defaut_lisse=pl.col("taux_defaut").cum_max()
        )
    if en_pourcent:
        cols = [c for c in ("taux_acceptation", "taux_defaut", "taux_defaut_lisse")
                if c in out.columns]
        out = out.with_columns([(pl.col(c) * 100).round(3).alias(c) for c in cols])

    noms = dict(COLONNES)
    if "taux_defaut_lisse" in out.columns:
        noms["taux_defaut_lisse"] = "Taux de défaut (lissé)"
    return out.select(list(noms)).rename(noms)


# ------------------------------------------------------------------- exports
def vers_csv(t: pl.DataFrame, francais: bool = True) -> bytes:
    """CSV. francais=True -> séparateur ';' et virgule décimale (Excel FR)."""
    csv = table_lisible(t).write_csv(separator=";" if francais else ",")
    if francais:
        csv = csv.replace(".", ",")          # décimales à la française
    return csv.encode("utf-8-sig")           # BOM : accents corrects dans Excel


def vers_excel(t: pl.DataFrame, seuil=None, feuille: str = "Seuils") -> bytes:
    """Classeur avec la table, les en-têtes figés et les formats de nombre."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    pdf = pd.DataFrame(table_lisible(t).to_dict(as_series=False))

    wb = Workbook()
    ws = wb.active
    ws.title = feuille
    ligne = 1

    if seuil is not None:
        ws.cell(row=1, column=1, value="Seuil retenu").font = Font(bold=True)
        ws.cell(row=1, column=2,
                value="tout accepté" if not seuil.atteint else float(seuil.seuil))
        ws.cell(row=2, column=1, value="Contrainte").font = Font(bold=True)
        ws.cell(row=2, column=2, value=seuil.contrainte)
        ws.cell(row=3, column=1, value="Acceptés").font = Font(bold=True)
        ws.cell(row=3, column=2, value=seuil.n_acceptes)
        ws.cell(row=3, column=3, value=seuil.taux_acceptation).number_format = "0.0%"
        ws.cell(row=4, column=1, value="Défaut des acceptés").font = Font(bold=True)
        ws.cell(row=4, column=2, value=seuil.taux_defaut).number_format = "0.00%"
        ligne = 6

    for j, nom in enumerate(pdf.columns, start=1):
        c = ws.cell(row=ligne, column=j, value=nom)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(pdf.itertuples(index=False), start=ligne + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            nom = pdf.columns[j - 1]
            if "Taux" in nom or "Part" in nom:
                cell.number_format = '0.00"%"'
            elif nom == "Seuil (proba)":
                cell.number_format = "0.000000"
            else:
                cell.number_format = "#,##0"

    ws.freeze_panes = ws.cell(row=ligne + 1, column=1)
    for j in range(1, len(pdf.columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ Streamlit
def afficher_tableau(st, t: pl.DataFrame, seuil=None, hauteur: int = 420,
                     autour: int | None = None):
    """
    Affiche la table des seuils + les deux boutons de téléchargement.

    autour : si renseigné, n'affiche que les N lignes autour du seuil retenu
             (la table complète reste téléchargeable).
    """
    vue = t
    if autour and seuil is not None and seuil.atteint:
        idx = (t.with_row_index("i")
                 .filter(pl.col("proba") <= seuil.seuil)["i"].max() or 0)
        vue = t.slice(max(0, idx - autour), 2 * autour + 1)

    pdf = pd.DataFrame(table_lisible(vue).to_dict(as_series=False))

    st.dataframe(
        pdf,
        use_container_width=True,
        hide_index=True,
        height=hauteur,
        column_config={
            "Seuil (proba)": st.column_config.NumberColumn(format="%.6f"),
            "Dossiers au niveau": st.column_config.NumberColumn(format="%d"),
            "Défauts au niveau": st.column_config.NumberColumn(format="%d"),
            "Acceptés (cumul)": st.column_config.NumberColumn(format="%d"),
            "Défauts acceptés (cumul)": st.column_config.NumberColumn(format="%d"),
            "Part d'acceptés": st.column_config.NumberColumn(format="%.2f %%"),
            "Taux de défaut des acceptés":
                st.column_config.NumberColumn(format="%.2f %%"),
            "Taux de défaut (lissé)":
                st.column_config.NumberColumn(format="%.2f %%"),
        },
    )
    st.caption(f"{t.height:,} niveaux de score".replace(",", " ")
               + (f" — vue centrée sur le seuil" if vue.height < t.height else ""))

    c1, c2 = st.columns(2)
    c1.download_button(
        "⬇️ CSV (table complète)",
        data=vers_csv(t),
        file_name="table_seuils.csv",
        mime="text/csv",
        use_container_width=True,
    )
    c2.download_button(
        "⬇️ Excel (table complète)",
        data=vers_excel(t, seuil),
        file_name="table_seuils.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    import numpy as np

    from seuil_octroi import seuil_pour_defaut, table_seuils

    rng = np.random.default_rng(3)
    n, k = 200_000, 200
    score = rng.integers(0, k, n)
    p = np.clip((score / k) ** 2 * .35 + .003, 0, 1)
    df = pl.DataFrame({"proba": np.round(p, 6), "defaut": rng.binomial(1, p)})

    t = table_seuils(df)
    s = seuil_pour_defaut(None, 0.05, table=t)
    print(table_lisible(t).head(4))
    open("table_seuils.xlsx", "wb").write(vers_excel(t, s))
    open("table_seuils.csv", "wb").write(vers_csv(t))
    print("exports écrits :", s)
