"""
Affichage Streamlit + export Excel des matrices de transition flag1 x flag2.

Dépend de matrice_transition.py (croise / en_pourcentage / avec_marges / ORDRE).
Les couleurs à l'écran et dans le fichier Excel sont calculées par la MÊME
fonction, donc le rendu est identique dans les deux.
"""

from io import BytesIO

import matplotlib
import pandas as pd
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from matrice_transition import ORDRE, avec_marges, croise, en_pourcentage

# ----------------------------------------------------------------- couleurs
GRIS_TOTAL = "EEEEEE"
BORDURE = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _to_pd(m: pl.DataFrame, f1: str = "flag1") -> pd.DataFrame:
    """Polars -> pandas indexé (sans passer par pyarrow)."""
    return pd.DataFrame(m.to_dict(as_series=False)).set_index(f1)


def _hex(cmap_name: str, x: float) -> str:
    """x dans [0,1] -> 'RRGGBB'. Plage 0.08-0.75 pour garder du texte lisible."""
    cmap = matplotlib.colormaps[cmap_name]
    r, g, b, _ = cmap(0.08 + 0.67 * max(0.0, min(1.0, x)))
    return f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"


def _echelle(pdf: pd.DataFrame, cols: list[str]) -> tuple[float, float]:
    bloc = pdf.loc[[i for i in pdf.index if i != "Total"], cols]
    return float(bloc.min().min()), float(bloc.max().max())


# ------------------------------------------------------------------ Streamlit
def styler(m: pl.DataFrame, fmt: str = "{:,.0f}", cmap: str = "Blues",
           f1: str = "flag1"):
    """Styler pandas prêt pour st.dataframe (dégradé + totaux en gras)."""
    pdf = _to_pd(m, f1)
    cols = [c for c in pdf.columns if c != "Total"]
    lo, hi = _echelle(pdf, cols)

    def couleur(val, ligne):
        if ligne == "Total":
            return f"background-color:#{GRIS_TOTAL};font-weight:600;"
        x = 0.5 if hi == lo else (val - lo) / (hi - lo)
        h = _hex(cmap, x)
        txt = "#FFFFFF" if x > 0.6 else "#111111"
        return f"background-color:#{h};color:{txt};"

    def appliquer(df):
        out = pd.DataFrame("", index=df.index, columns=df.columns)
        for i in df.index:
            for j in cols:
                out.loc[i, j] = couleur(df.loc[i, j], i)
            if "Total" in df.columns:
                out.loc[i, "Total"] = f"background-color:#{GRIS_TOTAL};font-weight:600;"
        return out

    return pdf.style.apply(appliquer, axis=None).format(fmt)


def afficher(st, m: pl.DataFrame, titre: str, fmt="{:,.0f}", cmap="Blues"):
    """Affiche une matrice colorée dans Streamlit."""
    st.markdown(f"**{titre}**")
    st.dataframe(styler(m, fmt=fmt, cmap=cmap), use_container_width=True)


# --------------------------------------------------------------------- Excel
def to_excel(blocs: list[dict], f1: str = "flag1", une_feuille: bool = True) -> bytes:
    """
    blocs = [{"titre":..., "df": pl.DataFrame, "cmap": "Blues",
              "fmt": "#,##0" ou '0.0"%"'}, ...]
    Renvoie les octets du .xlsx (à passer à st.download_button).
    """
    wb = Workbook()
    wb.remove(wb.active)
    if une_feuille:
        ws = wb.create_sheet("Matrices")
    ligne = 1

    for k, bloc in enumerate(blocs):
        pdf = _to_pd(bloc["df"], f1)
        cols = list(pdf.columns)
        num = [c for c in cols if c != "Total"]
        lo, hi = _echelle(pdf, num)
        cmap, fmt = bloc.get("cmap", "Blues"), bloc.get("fmt", "#,##0")

        if not une_feuille:
            ws = wb.create_sheet(bloc["titre"][:31].replace("/", "-"))
            ligne = 1

        # titre
        c = ws.cell(row=ligne, column=1, value=bloc["titre"])
        c.font = Font(bold=True, size=12)
        ligne += 1

        # en-tête
        ws.cell(row=ligne, column=1, value=f"{f1} \\ flag2").font = Font(bold=True)
        for j, nom in enumerate(cols, start=2):
            h = ws.cell(row=ligne, column=j, value=nom)
            h.font = Font(bold=True)
            h.alignment = Alignment(horizontal="center")
            h.fill = PatternFill("solid", fgColor="F2F2F2")
            h.border = BORDURE
        ligne += 1

        # corps
        for idx in pdf.index:
            r = ws.cell(row=ligne, column=1, value=idx)
            r.font = Font(bold=(idx == "Total"))
            r.fill = PatternFill("solid", fgColor="F2F2F2")
            r.border = BORDURE
            for j, nom in enumerate(cols, start=2):
                val = float(pdf.loc[idx, nom])
                cell = ws.cell(row=ligne, column=j, value=val)
                cell.number_format = fmt
                cell.border = BORDURE
                if idx == "Total" or nom == "Total":
                    cell.fill = PatternFill("solid", fgColor=GRIS_TOTAL)
                    cell.font = Font(bold=True)
                else:
                    x = 0.5 if hi == lo else (val - lo) / (hi - lo)
                    cell.fill = PatternFill("solid", fgColor=_hex(cmap, x))
                    if x > 0.6:
                        cell.font = Font(color="FFFFFF")
            ligne += 1
        ligne += 2  # respiration entre deux matrices

        ws.column_dimensions["A"].width = 16
        for j in range(2, len(cols) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------- construction des 4 blocs
def construire(df: pl.DataFrame, f1="flag1", f2="flag2", montant="montant",
               how="row", marges=True) -> list[dict]:
    """Les 4 matrices, prêtes à afficher et à exporter."""
    n = croise(df, f1, f2)
    m = croise(df, f1, f2, valeur=montant)
    n_pct, m_pct = en_pourcentage(n, f1, how), en_pourcentage(m, f1, how)
    if marges:
        n, m = avec_marges(n, f1), avec_marges(m, f1)
        if how == "all":                      # les % totalisent 100, marges utiles
            n_pct, m_pct = avec_marges(n_pct, f1), avec_marges(m_pct, f1)
    return [
        {"titre": "1. Effectifs (n)", "df": n, "cmap": "Blues", "fmt": "#,##0"},
        {"titre": f"2. % des effectifs ({how})", "df": n_pct, "cmap": "Purples",
         "fmt": '0.0"%"'},
        {"titre": "3. Montants sommés", "df": m, "cmap": "Greens",
         "fmt": "#,##0 €"},
        {"titre": f"4. % des montants ({how})", "df": m_pct, "cmap": "Oranges",
         "fmt": '0.0"%"'},
    ]


# ------------------------------------------------------- page Streamlit type
def page(st, df: pl.DataFrame):
    """À appeler depuis ta page existante : page(st, mon_df)."""
    c1, c2 = st.columns(2)
    how = c1.radio("Base des pourcentages", ["row", "col", "all"], horizontal=True,
                   format_func={"row": "par ligne", "col": "par colonne",
                                "all": "sur le total"}.get)
    marges = c2.checkbox("Afficher les totaux", value=True)

    blocs = construire(df, how=how, marges=marges)

    fmt_ecran = {"#,##0": "{:,.0f}", "#,##0 €": "{:,.0f} €", '0.0"%"': "{:.1f} %"}
    for b in blocs:
        afficher(st, b["df"], b["titre"], fmt=fmt_ecran[b["fmt"]], cmap=b["cmap"])

    st.download_button(
        "⬇️ Exporter en Excel",
        data=to_excel(blocs),
        file_name="matrices_transition.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import numpy as np
    rng = np.random.default_rng(0)
    n = 5000
    df = pl.DataFrame({
        "flag1": rng.choice(ORDRE, n, p=[.6, .25, .15]),
        "flag2": rng.choice(ORDRE, n, p=[.5, .3, .2]),
        "montant": rng.gamma(2, 5000, n).round(2),
    })
    with open("matrices_transition.xlsx", "wb") as f:
        f.write(to_excel(construire(df)))
    print("xlsx écrit")
